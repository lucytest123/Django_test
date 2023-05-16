import os.path
import random

import openpyxl
import xlrd
from openpyxl import load_workbook

identifier = 10000


# ecoding 测试数据导入脚本
def file_path(file_name="测试导入.xlsx"):
    file = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    files_path = os.path.join(file, "file")
    if not os.path.exists(files_path):
        os.makedirs(files_path)
    file_path = os.path.join(files_path, file_name)
    return file_path


def xls_file():
    print("请输入AE需要创造的行数")
    AE_number = int(input())
    print("请输入MH需要创造的行数")
    MH_number = int(input())
    print("请输入CE需要创造的行数")
    CE_number = int(input())
    print("请输入CM需要创造的行数")
    CM_number = int(input())
    print("是否创建用药信息 yes/no")
    Medication = input()
    # 写入文件1
    delect_file()
    xls_AE(AE_number)
    xls_CE(CE_number)
    xls_MH(MH_number)
    xls_CM(CM_number, Medication)
    # 写入AEsheet页面


def xls_AE(number):
    global identifier, sheel_list, i
    Verbatim = open_xls(0)
    if os.path.exists(file_path()):
        wookbooks = load_workbook(file_path())
    else:
        wookbooks = openpyxl.Workbook()
        wookbooks.save("测试导入.xlsx", )
    sheel = wookbooks.create_sheet("AE")
    wookbooks.remove(wookbooks["Sheet"])
    list = ["受试者标识符", "行号", "逐字术语"]
    sheel.append(list)
    i = 0
    while i < number:
        i = i + 1
        Line_number = random.randint(0, 99)
        identifier = identifier + 1
        sheel_list = [identifier, Line_number, Verbatim[random.randint(0, 300)]]

        sheel.append(sheel_list)
    wookbooks.save(file_path())
    wookbooks.close()


def delect_file():
    if os.path.exists(file_path()):
        os.remove(file_path())


def xls_MH(number):
    global identifier, sheel_list
    Verbatim = open_xls(0)
    if os.path.exists(file_path()):
        wookbooks = load_workbook(file_path())
    else:
        wookbooks = openpyxl.Workbook()
        wookbooks.save("测试导入.xlsx", )
    sheel = wookbooks.create_sheet("MH")
    list = ["受试者标识符", "行号", "逐字术语"]
    sheel.append(list)
    i = 0
    while i < number:
        i = i + 1
        Line_number = random.randint(0, 99)
        identifier = identifier + 1
        sheel_list = [identifier, Line_number, Verbatim[random.randint(0, 300)]]

        sheel.append(sheel_list)
    wookbooks.save(file_path())
    wookbooks.close()


def xls_CE(number):
    global identifier, i
    Verbatim = open_xls(0)
    if os.path.exists(file_path()):
        wookbooks = load_workbook(file_path())
    else:
        wookbooks = openpyxl.Workbook()
        wookbooks.save("测试导入.xlsx", )
    sheel = wookbooks.create_sheet("CE")
    list = ["受试者标识符", "行号", "逐字术语"]
    sheel.append(list)
    i = 0
    while i < number:
        i = i + 1
        Line_number = random.randint(0, 99)
        identifier = identifier + 1
        sheel_list = [identifier, Line_number, Verbatim[random.randint(0, 300)]]

        sheel.append(sheel_list)
    wookbooks.save(file_path())
    wookbooks.close()


def xls_CM(number, Medication):
    global identifier
    Verbatim = open_xls(0)
    Verbatims = open_xls(1)
    i = 0
    wookbooks = load_workbook(file_path())
    sheel = wookbooks.create_sheet("CM")
    list = ["受试者标识符", "行号", "逐字术语", "用药信息1", "用药信息2", "用药信息3"]
    sheel.append(list)
    while i < number:
        i = i + 1
        Line_number = random.randint(0, 99)
        identifier = identifier + 1
        if Medication != "yes":
            stu_list = [identifier, Line_number, Verbatim[random.randint(0, 300)]]
        else:
            stu_list = [identifier, Line_number, Verbatim[random.randint(0, 300)], Verbatims[random.randint(0, 500)],
                        Verbatims[random.randint(0, 500)],
                        Verbatims[random.randint(0, 500)]]

        sheel.append(stu_list)

    wookbooks.save(file_path())
    wookbooks.close()


#  sheets : 0 逐字术语 1 用药信息
def open_xls(sheets):
    xls_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    xls_file = os.path.join(xls_path, "file", "测试数据.xls")
    wookbooks = xlrd.open_workbook(xls_file)
    sheet = wookbooks.sheets()[int(sheets)]
    text = []
    for l in range(int(sheet.nrows)):
        x = sheet.row_values(l)
        for x in x:
            if "\n" in x:
                x.replace("\n", "")
            if "\r" in x:
                x.replace("\r", "")
            if "  " in x:
                x.replace("  ", "")
        text.append(x)
    return text


if __name__ == '__main__':
    xls_file()
